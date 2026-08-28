from __future__ import annotations

import csv
import io
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.domain.finance.gl_csv_validator import (
    REQUIRED_GL_COLUMNS,
    STANDARD_GL_COLUMNS,
    build_column_mapping,
)

SUPPORTED_GL_EXTENSIONS = {".csv", ".xlsx"}
XLSX_MAGIC = b"PK\x03\x04"
MAX_HEADER_SCAN_ROWS = 50


class UnsupportedGLFile(ValueError):
    pass


def extension_for(filename: str | None) -> str:
    return Path(filename or "general-ledger.csv").suffix.lower()


def ensure_supported_gl_filename(filename: str | None) -> str:
    value = Path(filename or "general-ledger.csv").name
    extension = extension_for(value)
    if extension not in SUPPORTED_GL_EXTENSIONS:
        raise UnsupportedGLFile(
            "FinCruiz supports CSV and Excel .xlsx files for General Ledger uploads. "
            "For legacy .xls files, save or export the workbook as .xlsx or CSV first."
        )
    return value


def _display_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value)


def _header_score(values: Iterable[Any]) -> tuple[int, int, int]:
    headers = [_display_cell(value).strip() for value in values]
    headers = [header for header in headers if header]
    if not headers:
        return (-1, -1, -1)

    mapping, details, issues = build_column_mapping(headers)
    detected = set(mapping.values())
    required = len(REQUIRED_GL_COLUMNS & detected)
    recognised = sum(1 for detail in details if detail.mapped_column in STANDARD_GL_COLUMNS and detail.method != "unmapped")
    hard_errors = sum(1 for issue in issues if issue.severity == "error")
    return (required, recognised, -hard_errors)


def _find_excel_sheet_and_header(workbook) -> tuple[str, int]:
    best_sheet: str | None = None
    best_row = 1
    best_score = (-1, -1, -1)
    first_nonempty: tuple[str, int] | None = None

    for worksheet in workbook.worksheets:
        if worksheet.sheet_state != "visible":
            continue
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=MAX_HEADER_SCAN_ROWS, values_only=True),
            start=1,
        ):
            values = list(row)
            if first_nonempty is None and any(_display_cell(value).strip() for value in values):
                first_nonempty = (worksheet.title, row_number)
            score = _header_score(values)
            if score > best_score:
                best_score = score
                best_sheet = worksheet.title
                best_row = row_number
            # All four mandatory GL fields were found. This is already a strong header.
            if score[0] == len(REQUIRED_GL_COLUMNS) and score[2] == 0:
                return worksheet.title, row_number

    if best_sheet is not None and best_score[0] >= 2:
        return best_sheet, best_row
    if first_nonempty is not None:
        return first_nonempty
    raise ValueError("The Excel workbook does not contain any data.")


def _open_xlsx(source: bytes | Path):
    try:
        if isinstance(source, Path):
            return load_workbook(filename=source, read_only=True, data_only=True)
        return load_workbook(filename=io.BytesIO(source), read_only=True, data_only=True)
    except (InvalidFileException, OSError, ValueError) as exc:
        raise ValueError(
            "The Excel workbook could not be read. Please upload a valid .xlsx file or export the General Ledger as CSV."
        ) from exc


def xlsx_bytes_to_csv_bytes(content: bytes) -> bytes:
    if not content.startswith(XLSX_MAGIC):
        raise ValueError("The selected .xlsx file is not a valid Excel workbook.")

    workbook = _open_xlsx(content)
    try:
        sheet_name, header_row = _find_excel_sheet_and_header(workbook)
    finally:
        workbook.close()

    workbook = _open_xlsx(content)
    try:
        worksheet = workbook[sheet_name]
        output = io.StringIO(newline="")
        writer = csv.writer(output)
        for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if row_number < header_row:
                continue
            values = [_display_cell(value) for value in row]
            if not any(value.strip() for value in values):
                continue
            writer.writerow(values)
        return output.getvalue().encode("utf-8-sig")
    finally:
        workbook.close()


def xlsx_path_to_csv_path(source: Path) -> Path:
    with source.open("rb") as handle:
        if handle.read(4) != XLSX_MAGIC:
            raise ValueError("The selected .xlsx file is not a valid Excel workbook.")

    workbook = _open_xlsx(source)
    try:
        sheet_name, header_row = _find_excel_sheet_and_header(workbook)
    finally:
        workbook.close()

    target = source.with_suffix(source.suffix + ".normalized.csv")
    workbook = _open_xlsx(source)
    try:
        worksheet = workbook[sheet_name]
        with target.open("w", encoding="utf-8-sig", newline="") as output:
            writer = csv.writer(output)
            for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                if row_number < header_row:
                    continue
                values = [_display_cell(value) for value in row]
                if not any(value.strip() for value in values):
                    continue
                writer.writerow(values)
    except Exception:
        target.unlink(missing_ok=True)
        raise
    finally:
        workbook.close()
    return target


def normalise_gl_bytes(filename: str | None, content: bytes) -> bytes:
    name = ensure_supported_gl_filename(filename)
    extension = extension_for(name)
    if extension == ".xlsx":
        return xlsx_bytes_to_csv_bytes(content)

    if content.startswith(XLSX_MAGIC):
        raise ValueError(
            "This file contains an Excel workbook but is named as CSV. "
            "Upload the original .xlsx file or export it as a real CSV file."
        )
    return content

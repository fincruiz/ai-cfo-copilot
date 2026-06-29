from io import BytesIO
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = df.columns.astype(str).str.strip()
    return df


def safe_float(value, default=0.0):
    try:
        if pd.isna(value):
            return default
        return float(value)
    except Exception:
        return default


def format_number(value):
    try:
        return f"{float(value):,.2f}"
    except Exception:
        return value


def round_numeric_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for col in df.select_dtypes(include="number").columns:
        df[col] = df[col].round(2)
    return df


def style_dataframe(df: pd.DataFrame):
    return df.style.format(precision=2).set_properties(**{
        "font-family": "Arial",
        "font-size": "13px",
        "text-align": "left"
    })


def format_excel_sheet(ws):
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(name="Arial", size=11, bold=True)
    body_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border

    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            max_length = max(max_length, len(str(cell.value)) if cell.value is not None else 0)
        ws.column_dimensions[col_letter].width = min(max_length + 3, 35)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22


def dataframe_to_excel_bytes(df_dict: dict[str, pd.DataFrame]) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in df_dict.items():
            safe_sheet = str(sheet_name)[:31]
            df.to_excel(writer, sheet_name=safe_sheet, index=False)
            format_excel_sheet(writer.book[safe_sheet])
    return output.getvalue()

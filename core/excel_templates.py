from io import BytesIO
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

def format_excel_sheet(ws):
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(name="Arial", size=11, bold=True)
    body_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                max_length = max(max_length, len(str(cell.value)) if cell.value is not None else 0)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 3, 40)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22


def dataframe_to_excel_bytes(df_dict: dict[str, pd.DataFrame]) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in df_dict.items():
            safe_sheet = str(sheet_name)[:31]
            if df is None:
                df = pd.DataFrame()
            df.to_excel(writer, sheet_name=safe_sheet, index=False)
            format_excel_sheet(writer.book[safe_sheet])
    return output.getvalue()


def make_sample_template_bytes(df: pd.DataFrame) -> bytes:
    return dataframe_to_excel_bytes({"Template": df})


def get_sample_templates():
    templates = {}
    templates["Current GL Report"] = pd.DataFrame([
        {"Account code": "4000", "Debit": 0, "Credit": 25000, "Branch": "Sydney", "Net": -25000, "Date": "2026-04-01", "Period": "April 2026", "Description": "Sales invoice"},
        {"Account code": "5100", "Debit": 8000, "Credit": 0, "Branch": "Sydney", "Net": 8000, "Date": "2026-04-02", "Period": "April 2026", "Description": "Freight domestic cost"},
        {"Account code": "5200", "Debit": 3000, "Credit": 0, "Branch": "Melbourne", "Net": 3000, "Date": "2026-04-03", "Period": "April 2026", "Description": "Freight international overhead"},
    ])
    templates["COA Mapping"] = pd.DataFrame([
        {"Account code": "4000", "Account Name": "Sales Revenue", "Reporting Group": "Revenue", "Reporting Subgroup": "Sales", "Statement": "Income Statement", "Sign Convention": "positive", "Display Order": 1},
        {"Account code": "5100", "Account Name": "Freight Domestic", "Reporting Group": "Cost of Sales", "Reporting Subgroup": "Freight Domestic", "Statement": "Income Statement", "Sign Convention": "positive", "Display Order": 2},
        {"Account code": "5200", "Account Name": "Freight International", "Reporting Group": "Operating Expense", "Reporting Subgroup": "Freight International", "Statement": "Income Statement", "Sign Convention": "positive", "Display Order": 4},
    ])
    templates["KPI Master"] = pd.DataFrame([
        {"KPI Name": "Revenue", "Formula Type": "direct", "Numerator Group": "Revenue", "Denominator Group": "", "Output Type": "value", "Display Order": 1},
        {"KPI Name": "COGS", "Formula Type": "direct", "Numerator Group": "Cost of Sales", "Denominator Group": "", "Output Type": "value", "Display Order": 2},
        {"KPI Name": "Gross Profit", "Formula Type": "derived", "Numerator Group": "Revenue", "Denominator Group": "Cost of Sales", "Output Type": "value", "Display Order": 3},
        {"KPI Name": "Gross Margin %", "Formula Type": "ratio", "Numerator Group": "Gross Profit", "Denominator Group": "Revenue", "Output Type": "percent", "Display Order": 4},
        {"KPI Name": "Operating Expenses", "Formula Type": "direct", "Numerator Group": "Operating Expense", "Denominator Group": "", "Output Type": "value", "Display Order": 5},
        {"KPI Name": "Operating Profit", "Formula Type": "derived", "Numerator Group": "Gross Profit", "Denominator Group": "Operating Expense", "Output Type": "value", "Display Order": 6},
        {"KPI Name": "Operating Margin %", "Formula Type": "ratio", "Numerator Group": "Operating Profit", "Denominator Group": "Revenue", "Output Type": "percent", "Display Order": 7},
        {"KPI Name": "Opex as % of Revenue", "Formula Type": "ratio", "Numerator Group": "Operating Expense", "Denominator Group": "Revenue", "Output Type": "percent", "Display Order": 8},
    ])
    templates["Latest Previous Balance Sheet"] = pd.DataFrame([
        {"Reporting Group": "Assets", "Reporting Subgroup": "Cash", "Balance": 50000},
        {"Reporting Group": "Liabilities", "Reporting Subgroup": "Trade Payables", "Balance": 22000},
        {"Reporting Group": "Equity", "Reporting Subgroup": "Retained Earnings", "Balance": 28000},
    ])
    templates["Budget Data"] = pd.DataFrame([
        {"Month": "2026-01", "Branch": "Sydney", "Reporting Group": "Revenue", "Amount": 100000},
        {"Month": "2026-01", "Branch": "Sydney", "Reporting Group": "Cost of Sales", "Amount": 60000},
        {"Month": "2026-01", "Branch": "Melbourne", "Reporting Group": "Revenue", "Amount": 85000},
    ])
    templates["Forecast P&L"] = pd.DataFrame([
        {"Period": "April 2026", "Reporting Group": "Revenue", "Reporting Subgroup": "Sales", "Report Value": 120000},
        {"Period": "April 2026", "Reporting Group": "Cost of Sales", "Reporting Subgroup": "Cost of Sales", "Report Value": 72000},
        {"Period": "April 2026", "Reporting Group": "Operating Expense", "Reporting Subgroup": "Rent", "Report Value": 15000},
    ])
    templates["Forecast Balance Sheet"] = pd.DataFrame([
        {"Reporting Group": "Assets", "Reporting Subgroup": "Cash", "Balance": 65000},
        {"Reporting Group": "Liabilities", "Reporting Subgroup": "Trade Payables", "Balance": 28000},
        {"Reporting Group": "Equity", "Reporting Subgroup": "Retained Earnings", "Balance": 37000},
    ])
    templates["Previous Year P&L"] = pd.DataFrame([
        {"Period": "April 2025", "Reporting Group": "Revenue", "Reporting Subgroup": "Sales", "Report Value": 98000},
        {"Period": "April 2025", "Reporting Group": "Cost of Sales", "Reporting Subgroup": "Cost of Sales", "Report Value": 59000},
        {"Period": "April 2025", "Reporting Group": "Operating Expense", "Reporting Subgroup": "Rent", "Report Value": 13000},
    ])
    templates["AR Ageing"] = pd.DataFrame([
        {"Party Name": "Customer A", "Outstanding Amount": 12000, "Document Number": "INV001", "Document Date": "2026-02-01", "Due Date": "2026-03-01", "Branch": "Sydney", "Age Bucket": "1-30"},
        {"Party Name": "Customer B", "Outstanding Amount": 8000, "Document Number": "INV002", "Document Date": "2026-01-15", "Due Date": "2026-02-15", "Branch": "Melbourne", "Age Bucket": "31-60"},
        {"Party Name": "Customer C", "Outstanding Amount": 5000, "Document Number": "INV003", "Document Date": "2026-03-05", "Due Date": "2026-04-05", "Branch": "Sydney", "Age Bucket": "Current"},
    ])
    templates["AP Ageing"] = pd.DataFrame([
        {"Party Name": "Supplier A", "Outstanding Amount": 9000, "Document Number": "BILL001", "Document Date": "2026-02-01", "Due Date": "2026-03-01", "Branch": "Sydney", "Age Bucket": "1-30"},
        {"Party Name": "Supplier B", "Outstanding Amount": 14000, "Document Number": "BILL002", "Document Date": "2026-01-10", "Due Date": "2026-02-10", "Branch": "Melbourne", "Age Bucket": "31-60"},
        {"Party Name": "Supplier C", "Outstanding Amount": 6000, "Document Number": "BILL003", "Document Date": "2026-03-04", "Due Date": "2026-04-04", "Branch": "Sydney", "Age Bucket": "Current"},
    ])
    templates["Industry Benchmark File"] = pd.DataFrame([
        {"Metric": "Gross Margin %", "Benchmark Value": 35},
        {"Metric": "Operating Margin %", "Benchmark Value": 12},
        {"Metric": "Opex as % of Revenue", "Benchmark Value": 20},
    ])
    templates["Prior Period P&L"] = templates["Previous Year P&L"].copy()
    templates["Prior Period Balance Sheet"] = templates["Latest Previous Balance Sheet"].copy()
    templates["Prior Period KPI Pack"] = pd.DataFrame([
        {"KPI": "Revenue", "Value": 98000, "Display Value": 98000, "Output Type": "value"},
        {"KPI": "Gross Margin %", "Value": 39.80, "Display Value": "39.80%", "Output Type": "percent"},
        {"KPI": "Operating Margin %", "Value": 26.53, "Display Value": "26.53%", "Output Type": "percent"},
    ])
    return templates

